import os, sys
import logging
from pathlib import Path
import config
import mysql.connector
from datetime import datetime
import time

CHECK_INTERVAL_SECONDS = getattr(config, "SLOTFILE_CHECK_INTERVAL", 5)
SLOTFILE_NAME = getattr(config, "SLOTFILE_NAME", "格口方案.xlsx")
PROCESSED_DIR = getattr(config, "SLOTFILE_PROCESSED_DIR", "processed")

# def read_slot_excel(file_path):
#     file_path = Path(file_path)
#     if not file_path.exists():
#         logging.info("不存在格口方案表: %s",str(file_path))
#     try:
#         from openpyxl import load_workbook
#         wb = load_workbook(filename=str(file_path),read_only=True, data_only=True)
#         ws = wb.active
#         rows = ws.iter_cols(values_only=True)
#         header = None
#         data_rows = []
#         for i,row in enumerate(rows):
#             if row is None:
#                 continue
#             row_list = list(row)
#             if i == 0:
#                 header = [(str(x).strip() if x is not None else "") for x in row_list]
#                 continue
#             data_rows.append([(str(x).strip() if x is not None else "") for x in row_list])
        
#         col_map = {}
#         if header:
#             for idx,colname in enumerate(header):
#                 low = colname.lower()
#                 if "一段" in colname or "一段码" in colname:
#                     col_map["terminal"] = idx
#                 if "格口" in colname or "格口号" in colname:
#                     col_map["slot"] = idx
#         if 'terminal' not in col_map or 'slot' not in col_map:
#             col_map['terminal'] = 0
#             col_map['slot'] = 1
#         result = []
#         for r in data_rows:
#             t_idx = col_map.get('terminal',0)
#             s_idx = col_map.get('slot',1)
#             term = r[t_idx] if t_idx<len(r) else ""
#             slot = r[s_idx] if s_idx<len(r) else ""
#             term = str(term).strip()
#             slot = str(slot).strip()
#             if term == "" and slot == "":
#                 continue
#             result.append((term,slot))
#         return result
#     except Exception as e_openpyxl:
#         logging.info("openpyxl read failed: %s, try pandas fallback",e_openpyxl)
#         try:
#             import pandas as pd
#             df = pd.read_excel(str(file_path), dtype=str)
#             if df.empty():
#                 return []
#             cols = list(df.columns)
#             terminal_col = None
#             slot_col = None
#             for c in cols:
#                 low = str(c).lower()
#                 if "一段" in str(c):
#                     terminal_col = c
#                 if "格口" in str(c):
#                     slot_col = c
#             if terminal_col is None or slot_col is None:
#                 terminal_col = cols[0]
#                 slot_col = cols[1] if len(cols) > 1 else cols[0]
#             pairs = []
#             for _, row in df.iterrows():
#                 t = str(row.get(terminal_col,"")).strip()
#                 s = str(row.get(slot_col,"")).strip()
#                 if t=="" and s == "":
#                     continue
#                 pairs.append((t,s))
#             return pairs
#         except Exception as e_pd:
#             logging.info("读取excel失败： openpyxl 和pandas 读取失败： %s, %s",e_openpyxl, e_pd)
def read_slot_excel(file_path):
    """
    读取 Excel 文件，返回 list of (terminal_code, slot_id).
    优先使用 openpyxl（不使用 read_only 模式），回退使用 pandas。
    始终返回列表（可能为空），出错会抛异常（上层捕获并记录）。
    """
    from pathlib import Path
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(str(file_path))

    # 1) 尝试 openpyxl（不使用 read_only，这样 sheet 有更多方法）
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(file_path), read_only=False, data_only=True)
        ws = wb.active

        # 迭代所有行（values_only=True 返回 tuple）
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        # 找到第一行不是全空的那一行作为 header_candidate（可能是 header，也可能就是数据）
        header_row = None
        data_start_index = 0
        for i, row in enumerate(rows):
            if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                header_row = [str(cell).strip() if cell is not None else "" for cell in row]
                data_start_index = i + 1
                break

        # 如果 header_row 看起来像“列名”（包含中文关键字或非数字），认为它是 header；否则把它当作数据
        use_header = False
        if header_row:
            joined = " ".join(header_row).lower()
            # 常见 header 关键词，按需扩展
            if any(k in joined for k in ["一段", "格口", "waybill", "slot", "格口号", "waybillno"]):
                use_header = True
            else:
                # 如果 header_row 中存在非纯数字的值，也可以认为是 header（保守判断）
                non_numeric_count = sum(0 if (cell is None or str(cell).strip() == "") else (1 if not str(cell).strip().replace('.','',1).isdigit() else 0) for cell in header_row)
                if non_numeric_count >= 1:
                    use_header = True

        # 如果确定 header 存在，则从 data_start_index 开始是数据；否则把 header_row 也作为数据，从 data_start_index-1 开始
        results = []
        if use_header:
            # map header name -> index
            header = header_row
            term_idx = None
            slot_idx = None
            for idx, colname in enumerate(header):
                low = colname.lower()
                if "一段" in colname or "一段码" in colname or "waybill" in low or "waybillno" in low:
                    term_idx = idx
                if "格口" in colname or "格口号" in colname or "slot" in low or "slot_id" in low:
                    slot_idx = idx
            # fallback to first two columns
            if term_idx is None: term_idx = 0
            if slot_idx is None: slot_idx = 1 if len(header) > 1 else 0

            for r in rows[data_start_index:]:
                if not r:
                    continue
                r_list = [ (str(c).strip() if c is not None else "") for c in r ]
                term = r_list[term_idx] if term_idx < len(r_list) else ""
                slot = r_list[slot_idx] if slot_idx < len(r_list) else ""
                if term == "" and slot == "":
                    continue
                results.append((term, slot))
        else:
            # treat all rows as data, use first two columns
            for r in rows:
                if not r:
                    continue
                r_list = [ (str(c).strip() if c is not None else "") for c in r ]
                # skip completely empty rows
                if all(x == "" for x in r_list):
                    continue
                term = r_list[0] if len(r_list) > 0 else ""
                slot = r_list[1] if len(r_list) > 1 else ""
                if term == "" and slot == "":
                    continue
                results.append((term, slot))

        return results
    except Exception as e_open:
        logging.debug("openpyxl read failed: %s, try pandas fallback", e_open)

    # 2) pandas fallback（更鲁棒）
    try:
        import pandas as pd
        df = pd.read_excel(str(file_path), engine="openpyxl", dtype=str)  # 读成字符串，避免类型问题
        if df is None or df.empty:
            return []
        # 丢弃全为空行
        df = df.dropna(how='all')
        if df.empty:
            return []
        cols = list(df.columns)
        # find terminal and slot columns
        terminal_col = None
        slot_col = None
        for c in cols:
            low = str(c).lower()
            if "一段" in str(c) or "waybill" in low or "waybillno" in low:
                terminal_col = c
            if "格口" in str(c) or "slot" in low or "slot_id" in low:
                slot_col = c
        if terminal_col is None or slot_col is None:
            terminal_col = cols[0]
            slot_col = cols[1] if len(cols) > 1 else cols[0]

        pairs = []
        for _, row in df.iterrows():
            t = str(row.get(terminal_col, "")).strip()
            s = str(row.get(slot_col, "")).strip()
            if t == "" and s == "":
                continue
            pairs.append((t, s))
        return pairs
    except Exception as e_pd:
        logging.exception("读取excel失败： openpyxl 和 pandas 读取失败： %s, %s", e_open, e_pd)
        # 保证不返回 None，上层会捕获异常并保留文件以便下次重试
        raise

def update_terminal_to_slot_pairs(pairs):
    """
    使用mysql connector 连接DB ，清空terminal_to_slot 表并插入pairs
    """
    if not pairs:
        logging.info("slot file parsed 0 rows, skip DB update")
        return
    db_conf = getattr(config,"DB_CONFIG",None)
    if not db_conf:
        logging.info("DB_CONFIG not found in config")
        return
    conn = None
    try:
        conn = mysql.connector.connect(**db_conf)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM terminalcode_to_slot")
        conn.commit()
        logging.info("Cleared terminalcode_to_slot table")

        insert_sql = "INSERT INTO terminalcode_to_slot (terminal_code, slot_id) VALUES (%s, %s)"
        to_insert = []
        for term,slot in pairs:
            slot_val = slot
            try:
                if slot_val is not None and slot_val !="":
                    slot_val = int(float(slot_val))
            except Exception:
                slot_val = slot
            to_insert.append((term,slot_val))
        cursor.executemany(insert_sql,to_insert)
        conn.commit()
        logging.info("Inserted %d rows into terminalcode_to_slot",len(to_insert))
    except Exception as e:
        logging.info("更新 terminalcode_to_slot 发生异常:%s",e)
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
    finally:
        if conn:
            try:
                cursor.close()
                conn.close()
            except Exception:
                pass
def slotfile_watcher_loop(stop_event):
    cwd = Path.cwd()
    processed_dir = cwd/PROCESSED_DIR
    processed_dir.mkdir(parents=True, exist_ok=True)
    logging.info("Slotfile watcher started: checking every %s seconds for %s ", CHECK_INTERVAL_SECONDS,SLOTFILE_NAME)

    while not stop_event.is_set():
        try:
            target = cwd / SLOTFILE_NAME
            if target.exists():
                logging.info("Found slot file : %s",str(target))
                try:
                    pairs = read_slot_excel(target)
                    logging.info("Parsed %d pairs from slot file",len(pairs))
                    update_terminal_to_slot_pairs(pairs)
                    now = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = processed_dir / f"{target.stem}_{now}{target.suffix}"
                    target.rename(dest)
                    logging.info("Moved processed slot file to %s",dest)
                except Exception as e:
                    logging.info("处理 slot 文件时出错, 将保留源文件以便下次重试: %s",e)
            for _ in range(int (CHECK_INTERVAL_SECONDS*10)):
                if stop_event.is_set():
                    break
                time.sleep(0.1)
        except Exception:
            logging.info("slot file wathcer loop 内部异常")
            time.sleep(CHECK_INTERVAL_SECONDS)
